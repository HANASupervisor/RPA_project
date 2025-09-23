from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook
wb = load_workbook('sample.xlsx')
ws = wb.active

# 번호, 수학, 영어
a1 = ws['A1'] # 번호
b1 = ws['B1'] # 수학
c1 = ws['C1'] # 영어

# A 열의 너비를 5로 설정
ws.column_dimensions['A'].width = 5
# 1행의 높이를 50으로 설정
ws.row_dimensions[1].height = 50

# 스타일 적용
a1.font = Font(color='FF0000', italic=True, bold=True) # 글자 색은 빨갛게, 이탤릭 두껍게
b1.font = Font(color='CC33FF', name='Arial', strike=True) # 폰트를 Arial로 설정, 취소선을 적용, strike는 취소선임.
c1.font = Font(color='0000FF', size = 20, underline='single') # 글자 크기는 20으로 밑줄을 적용한다. 

# 테두리 적용
thin_border = Border(left= Side(style='thin'), right = Side(style='thin'), top = Side(style='thin'), bottom = Side(style='thin'))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 만약 영어나 수학이 90점이 넘는경우 초록색으로 표시해서 어떤친구가 성적이 좋았는지를 표시할수 있게 하는 기능
for row in ws.rows:
    for cell in row:
        # 각 cell에대해 정렬
        cell.alignment = Alignment(horizontal='center', vertical='center')
        # center, left, right, top, bottom

        if cell.column == 1: # A는 학생번호니까 이거 제외
            continue
        
        # isinstance 쓰는이유는 이거 셀에 영어랑 수학이라는 문자형도 각 열에 하나씩들어갔기때문에 
        # 정수인경우만 계산하고 문자열인 경우 무시하고 나머지 계산을 하라는 함수 사용하기위해 isinstance함수 사용하는거임.
        if isinstance(cell.value, int) and cell.value > 90: # 만약 셀이 정수형 데이터이고 90점을 넘는경우
            cell.fill = PatternFill(fgColor='00FF00', fill_type='solid') # 배경을 초록색으로
            cell.font = Font(color='FF0000') # 폰트 색상 변경

# 틀 고정
ws.freeze_panes = 'B2' # B2 기준으로 윗줄과 아랫줄을 구분하는 굵은 선을 긋는다. 

wb.save('sample_style.xlsx')