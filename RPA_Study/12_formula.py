import datetime
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws['A1'] = datetime.datetime.today() # 오늘 날짜 저장
ws['A2'] = '=SUM(1,2,3)' # 엑셀 함수 자동 실행
ws['A3'] = '=AVERAGE(1,2,3)' # 2평균

ws['A4'] = 10
ws['A5'] = 20
ws['A6'] = '=SUM(A4:A5)' # 30 데이터 연산 가능함.

wb.save('sample_formula.xlsx')


