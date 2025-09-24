from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = 'project1'

# A1 cell에 1넣기
ws['A1'] = 1
ws['A2'] = 2
ws['A3'] = 3

ws['B1'] = 4
ws['B2'] = 5
ws['B3'] = 6

print(ws['A1']) # A1 cell 정보 출력
print(ws['A1'].value) # A1 cell의 값)
print(ws['A10'].value) # 값이 없을때 None을 출력


print(ws.cell(row=1, column=1).value)
print(ws.cell(row=1, column=2).value)

c = ws.cell(column=3, row=1, value=10) # ws['C1'].value = 10
print(c.value)

# 반복문통해 여러셀의 데이터를 한번에 조회
index = 1
from random import *
for x in range(1, 11): # 10개의 row
  for y in range(1, 11): # 10개의 col에대해 데이터를 채운다
    # ws.cell(row=x, column=y, value=randint(0,100))
    ws.cell(row=x, column=y, value=index)
    index += 1


































wb.save('sample.xlsx')
wb.close()

print('끝')