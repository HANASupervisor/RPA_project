from openpyxl import load_workbook
wb = load_workbook('sample_merged.xlsx')
ws = wb.active

# 지금 B2부터 D2까지 병합되어있음 이제 이거 나눌거임
ws.unmerge_cells('B2:D2')
wb.save('sample_unmerge.xlsx')
wb.close()
