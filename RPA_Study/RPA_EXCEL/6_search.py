#########################################
# 데이터 생성하는 코드
#########################################
# from openpyxl import load_workbook
# from openpyxl import *
# from random import *

# wb = Workbook()
# ws = wb.active

# # 열에대한 정보를 첫번째 열에 추가
# ws['A1'] = '번호'
# ws['B1'] = '수학'
# ws['C1'] = '영어'

# for row in range(2, 11):
#     ws[f'A{row}'] = row # 데이터 index번호 채우기
#     ws[f'B{row}'] = randint(0,100) # 랜덤하게 수학점수부여
#     ws[f'C{row}'] = randint(0,100) # 랜덤하게 영어점수부여


# wb.save('sample.xlsx')
# wb.close()


############################################
# 데이터 불러오기
############################################
from openpyxl import load_workbook
wb = load_workbook('sample.xlsx')
ws = wb.active

# 수학이 천재인 학생 뽑기
for row in ws.iter_rows(min_row =2):
    if row[2].value > 80:
        print(f'{row[0].value} 번 학생은 수학이 천재군요')


# 수학성적이아니라 컴퓨터 성적인 경우
for row in ws.iter_rows(max_row = 1):
    for cell in row:
        if cell.value == '수학':
            cell.value = '컴퓨터'

wb.save('sample_modify.xlsx')
print('작업 끝')

