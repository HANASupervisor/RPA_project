# Quiz) 여러분은 통계학과 교수님입니다. 
# 여러분이 가르치는 과목의 점수 비중은 다음과 같습니다. 

# - 출석: 10
# - 퀴즈1: 10
# - 퀴즈2: 10
# - 중간고사: 20
# - 기말고사: 30
# - 프로젝트: 20

# - 총합 100

# 마지막 수업을 모두 마치고 이번 학기 학생들의 최종 성적을 검토하는 과정에서 
# 퀴즈2 문제에 오류를 발견하여 모두 만점 처리를 하기로 하였습니다. 
# 현재까지 작성된 최종 성적 데이터를 기준으로 아래와 같이 수정하시오.

# 1. 퀴즈2 점수를 10으로 수정
# 2. H열에 총점(SUM 이용), I열에 성적 정보 추가
# - 총점 90 이상 A, 80 이상 B, 70 이상 C, 나머지 D
# 3. 출석이 5 미만인 학생은 총점 상관없이 F

# * 최종 파일명: scores.xlsx

from openpyxl import load_workbook
wb = load_workbook('problem1.xlsx')
ws = wb.active

# 퀴즈2 점수를 모두 10으로 수정
for i in range(1, ws.max_row + 1):
  if i != 1:
    ws[f'D{i}'].value = 10
  else:
     continue

# 2. H열에 총점(SUM 이용), I열에 성적 정보 추가
for i in range(1, ws.max_row + 1):
    ws[f'H{i}'] = f'=SUM(B{i}:G{i})'

# 3. I열에 성적 정보 추가
# - 총점 90 이상 A, 80 이상 B, 70 이상 C, 나머지 D
# 3. 출석이 5 미만인 학생은 총점 상관없이 F

# 일단 row_num이 1인 경우는 변수명이 들어있는 자리니까 제외하고 분석한다
for row_num in range(2, ws.max_row + 1):
   if ws[f'B{row_num}'].value < 5:
      ws[f'I{row_num}'] = 'F'


   # 성적을 매길건데 B{row_num}부터 G{row_num}까지 더한다. 
   total_score = ws[f'B{row_num}'].value + ws[f'C{row_num}'].value + ws[f'D{row_num}'].value + ws[f'E{row_num}'].value + ws[f'F{row_num}'].value + ws[f'G{row_num}'].value
      
   if total_score >= 90:
      ws[f'I{row_num}'].value = 'A'
      continue
   
   elif total_score >= 80:
      ws[f'I{row_num}'].value = 'B'
      continue
   
   elif total_score >= 70:
      ws[f'I{row_num}'].value = 'C'
      continue
   
   else:
      ws[f'I{row_num}'].value = 'D'
      continue
      
      

wb.save('scores.xlsx')