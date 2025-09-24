from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet() # 새로운 sheet 기본 이름으로 생성
ws.title = 'MySheet' # sheet 이름 변경
ws.sheet_properties.tabColor = 'ff66ff' # 탭 색상 변경

# 새로운 sheet한번 더 만들어본다.
ws1 = wb.create_sheet('YourSheet') # 주어진 이름으로 sheet 생성
ws2 = wb.create_sheet('NewSheet', 2) # 2번째 index에 sheet 생성

new_ws = wb['NewSheet'] # dict형태로 sheet에 접근

print(wb.sheetnames) # 모든 sheet 이름 확인

# sheet 복사 
new_ws['A1'] = 'Test'
target = wb.copy_worksheet(new_ws)
target.title = 'Copied Sheet'


wb.save('sample.xlsx')
wb.close()